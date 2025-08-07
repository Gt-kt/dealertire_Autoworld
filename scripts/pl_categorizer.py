# /scripts/pl_categorizer.py
import openpyxl
from openpyxl.styles import PatternFill
import io

# --- Configuration ---
# openpyxl uses ARGB hex codes for colors. FFFF00 is yellow.
NEW_VENDOR_COLOR = "FFFF00"
DEFAULT_CATEGORY = '공통'

def find_column_indices(sheet):
    """Finds the column index for '거래처명' and '구분' using openpyxl."""
    vendor_col, category_col = None, None
    if sheet.max_row < 1:
        return None, None
        
    # Read the first row to find headers
    headers = [cell.value for cell in sheet[1]]
    try:
        vendor_col = headers.index('거래처명') + 1
    except (ValueError, TypeError):
        pass
    try:
        category_col = headers.index('구분') + 1
    except (ValueError, TypeError):
        pass
        
    return vendor_col, category_col

def build_category_map(workbook):
    """Builds a category map from the previous month's workbook using openpyxl."""
    category_map = {}
    for sheet in workbook.worksheets:
        vendor_col, category_col = find_column_indices(sheet)

        if vendor_col and category_col:
            # Iterate from the second row to skip the header
            for row_idx in range(2, sheet.max_row + 1):
                vendor = sheet.cell(row=row_idx, column=vendor_col).value
                category = sheet.cell(row=row_idx, column=category_col).value
                
                if vendor and category:
                    vendor_str = str(vendor).strip()
                    if vendor_str not in category_map:
                        category_map[vendor_str] = str(category).strip()
    return category_map

def process_workbook(workbook, category_map):
    """Processes the current month's workbook using openpyxl."""
    # Define the highlight style once
    highlight_fill = PatternFill(start_color=NEW_VENDOR_COLOR, end_color=NEW_VENDOR_COLOR, fill_type="solid")
    
    for sheet in workbook.worksheets:
        vendor_col, category_col = find_column_indices(sheet)

        if not vendor_col:
            continue  # Skip sheet if no vendor column

        if not category_col:
            # Add a new '구분' column if it doesn't exist
            category_col = sheet.max_column + 1
            sheet.cell(row=1, column=category_col).value = '구분'

        # Process rows
        for row_idx in range(2, sheet.max_row + 1):
            vendor_cell = sheet.cell(row=row_idx, column=vendor_col)
            vendor_name = str(vendor_cell.value).strip() if vendor_cell.value else None

            if not vendor_name:
                continue

            category_cell = sheet.cell(row=row_idx, column=category_col)
            
            if vendor_name in category_map:
                category_cell.value = category_map[vendor_name]
            else:
                category_cell.value = DEFAULT_CATEGORY
                # Highlight the entire row for new vendors
                for cell in sheet[row_idx]:
                    cell.fill = highlight_fill
    
    # The modified workbook object is returned implicitly
    return workbook

def process_files(previous_file_stream, current_file_stream):
    """
    Main function to orchestrate building the map and processing the current file.
    Returns the processed file as an in-memory buffer.
    """
    # 1. Build the category map from the previous month's file
    wb_prev = openpyxl.load_workbook(previous_file_stream, read_only=True)
    category_map = build_category_map(wb_prev)
    
    if not category_map:
        raise ValueError("Could not build a category map from the 'previous month' file. Please check its format and content.")

    # 2. Process the current month's file using the map
    wb_curr = openpyxl.load_workbook(current_file_stream)
    processed_wb = process_workbook(wb_curr, category_map)
    
    # 3. Save the result to an in-memory buffer
    output_buffer = io.BytesIO()
    processed_wb.save(output_buffer)
    output_buffer.seek(0)
    
    return output_buffer