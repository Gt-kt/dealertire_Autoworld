# /scripts/pl_converter.py
import openpyxl
import re
import io
from openpyxl.utils import get_column_letter

# --- Configuration: All constants and helper functions are copied directly ---

def normalize_d1_name(name):
    """Cleans and normalizes account names from Dataset 1 for reliable lookup."""
    if not isinstance(name, str):
        return ""
    name = name.replace('\xa0', ' ')
    name = re.sub(r"^\s*\d+\.\s*", "", name)
    name = re.sub(r"\s*\([^)]*\)", "", name)
    name = re.sub(r"\s*\[[^\]]*\]", "", name)
    name = ' '.join(name.split())
    return name

SUMMABLE_NORMALIZED_KEYS = {
    "멤버십 구독료"
}

# This mapping should be defined at the top level so normalize_d1_name is available
MAP_D2_TO_D1 = [
    {"d2_display_name": "1. 매         출", "d1_lookup_names": [normalize_d1_name("1. 매 출")], "type": "direct"},
    {"d2_display_name": "  상품매출", "d1_lookup_names": [normalize_d1_name("상 품 매 출")], "type": "direct"},
    {"d2_display_name": "    B2B_타이어매출", "d1_lookup_names": [normalize_d1_name("B2B_타이어매출")], "type": "direct"},
    {"d2_display_name": "    B2B_부품매출", "d1_lookup_names": [normalize_d1_name("B2B_부품매출")], "type": "direct"},
    {"d2_display_name": "    B2C_타이어매출", "d1_lookup_names": [normalize_d1_name("B2C_타이어매출")], "type": "direct"},
    {"d2_display_name": "    B2C_부품매출", "d1_lookup_names": [normalize_d1_name("B2C_부품매출")], "type": "direct"},
    {"d2_display_name": "  용 역  매 출", "d1_lookup_names": [normalize_d1_name("용 역 매 출")], "type": "direct"},
    {"d2_display_name": "    B2C_용역매출", "d1_lookup_names": [normalize_d1_name("B2C_용역매출")], "type": "direct"},
    {"d2_display_name": "    멤버십 구독료", "d1_lookup_names": ["멤버십 구독료"], "type": "direct"},
    {"d2_display_name": "    기타 용역매출", "d1_lookup_names": [normalize_d1_name("기타 용역매출")], "type": "direct"},
    {"d2_display_name": "    Fleet 매출(쏘카)", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "2. 매  출   원  가", "d1_lookup_names": [normalize_d1_name("2. 매 출 원 가")], "type": "direct"},
    {"d2_display_name": "  상품매출원가", "d1_lookup_names": [normalize_d1_name("상품매출원가")], "type": "direct"},
    {"d2_display_name": "    타이어 매출원가", "d1_lookup_names": [normalize_d1_name("타이어 매출원가")], "type": "direct"},
    {"d2_display_name": "    부품 매출원가", "d1_lookup_names": [normalize_d1_name("부품 매출원가")], "type": "direct"},
    {"d2_display_name": "  용역매출원가(총)", "d1_lookup_names": [normalize_d1_name("용역매출원가")], "type": "direct"},
    {"d2_display_name": "    용역매출원가_쏘카", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "    B2C 용역매출원가", "d1_lookup_names": [normalize_d1_name("B2C 용역매출원가")], "type": "direct"},
    {"d2_display_name": "    기타용역매출원가", "d1_lookup_names": [normalize_d1_name("기타용역매출원가")], "type": "direct"},
    {"d2_display_name": "3. 매  출   총   이  익", "d1_lookup_names": [normalize_d1_name("3. 매 출 총 이 익")], "type": "direct"},
    {"d2_display_name": "4. 판매비 및  일반관리비", "d1_lookup_names": [normalize_d1_name("4. 판매비 및  일반관리비")], "type": "direct"},
    {"d2_display_name": "  급       여", "d1_lookup_names": [normalize_d1_name("급     여")], "type": "direct"},
    {"d2_display_name": "  잡   급", "d1_lookup_names": [normalize_d1_name("잡     급")], "type": "direct"},
    {"d2_display_name": "  퇴 직  급 여", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "  복 리 후 생 비", "d1_lookup_names": [normalize_d1_name("복 리 후 생 비")], "type": "direct"},
    {"d2_display_name": "  여 비  교  통  비", "d1_lookup_names": [normalize_d1_name("여 비 교 통 비")], "type": "direct"},
    {"d2_display_name": "  접   대   비", "d1_lookup_names": [normalize_d1_name("접   대   비")], "type": "direct"},
    {"d2_display_name": "  통      신      비", "d1_lookup_names": [normalize_d1_name("통     신   비")], "type": "direct"},
    {"d2_display_name": "  소 모  품 비", "d1_lookup_names": [normalize_d1_name("소 모 품 비")], "type": "direct"},
    {"d2_display_name": "  세 금 과 공 과", "d1_lookup_names": [normalize_d1_name("세 금 과 공 과")], "type": "direct"},
    {"d2_display_name": "  감 가  상  각  비", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "  지 급  임  차  료", "d1_lookup_names": [normalize_d1_name("지 급 임 차 료")], "type": "direct"},
    {"d2_display_name": "  수      선      비", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "  렌      탈      료", "d1_lookup_names": [normalize_d1_name("렌     탈   료")], "type": "direct"},
    {"d2_display_name": "  보  험  료", "d1_lookup_names": [normalize_d1_name("보   험   료")], "type": "direct"},
    {"d2_display_name": "  차 량 유 지 비", "d1_lookup_names": [normalize_d1_name("차 량 유 지 비")], "type": "direct"},
    {"d2_display_name": "  교 육 훈 련 비", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "  수 도  광  열  비", "d1_lookup_names": [normalize_d1_name("수 도 광 열 비")], "type": "direct"},
    {"d2_display_name": "  지 급 수 수 료", "d1_lookup_names": [normalize_d1_name("지 급 수 수 료")], "type": "direct"},
    {"d2_display_name": "    PG 지급수수료", "d1_lookup_names": [normalize_d1_name("PG 지급수수료")], "type": "direct"},
    {"d2_display_name": "    기타 지급수수료", "d1_lookup_names": [normalize_d1_name("지급수수료_소프트웨어"), normalize_d1_name("기타 지급수수료")], "type": "calculation", "op": "sum"},
    {"d2_display_name": "    지급수수료_위탁판매수수료", "d1_lookup_names": [normalize_d1_name("지급수수료_위탁판매수수료")], "type": "direct"},
    {"d2_display_name": "  도 서  인  쇄  비", "d1_lookup_names": [normalize_d1_name("도 서 인 쇄 비")], "type": "direct"},
    {"d2_display_name": "  외주용역비", "d1_lookup_names": [normalize_d1_name("외 주 용 역 비")], "type": "direct"},
    {"d2_display_name": "    외 주 용 역 비[쏘카]", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "    외주용역비[장착]", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "  광고선전비", "d1_lookup_names": [normalize_d1_name("광 고 선 전 비")], "type": "direct"},
    {"d2_display_name": "  건 물  관  리  비", "d1_lookup_names": [normalize_d1_name("건 물 관 리 비")], "type": "direct"},
    {"d2_display_name": "  운   반   비", "d1_lookup_names": [normalize_d1_name("운     반   비")], "type": "direct"},
    {"d2_display_name": "5. 영   업   손   익", "d1_lookup_names": [normalize_d1_name("5. 영 업 손 익")], "type": "direct"},
    {"d2_display_name": "6. 영   업   외   수   익", "d1_lookup_names": [normalize_d1_name("6. 영 업 외 수 익")], "type": "direct"},
    {"d2_display_name": "  이   자   수   익", "d1_lookup_names": [normalize_d1_name("이 자 수 익")], "type": "direct"},
    {"d2_display_name": "  수 입  임  대  료", "d1_lookup_names": [normalize_d1_name("수 입 임 대 료")], "type": "direct"},
    {"d2_display_name": "  외   환   차   익", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "  잡       이       익", "d1_lookup_names": [normalize_d1_name("잡     이     익")], "type": "direct"},
    {"d2_display_name": "7. 영   업   외   비   용", "d1_lookup_names": [normalize_d1_name("7. 영 업 외 비 용")], "type": "direct"},
    {"d2_display_name": "  이   자   비   용", "d1_lookup_names": [normalize_d1_name("이 자 비 용")], "type": "direct"},
    {"d2_display_name": "  외   환   차   손", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "  잡       손       실", "d1_lookup_names": [], "type": "blank"},
    {"d2_display_name": "8. 법인세비용차감전순손익", "d1_lookup_names": [normalize_d1_name("8. 법인세비용차감전순손익")], "type": "direct"},
    {"d2_display_name": "12. 당 기  순  이  익", "d1_lookup_names": [normalize_d1_name("9. 당 기 순 이 익")], "type": "direct"}
]

# --- Core Logic ---

def process_file(input_file):
    """
    Reads an Excel file stream, processes it, adds new sheets, and returns the result.
    """
    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet1 = workbook.active
        
        data_start_row = None
        data_end_row = None
        
        # Find the start of the data by looking for "계정명"
        for r_idx, row in enumerate(sheet1.iter_rows(min_row=1, max_col=1, values_only=True), 1):
            if isinstance(row[0], str) and "계정명" in row[0]:
                data_start_row = r_idx + 1
                break
        
        if not data_start_row:
            raise ValueError("Could not find the '계정명' header row in the input file.")

        # Find the end of the data more safely
        footer_pattern = re.compile(r"^\d{4}/\d{2}/\d{2}\s+(오전|오후)\s+\d{1,2}:\d{2}:\d{2}")
        for r_idx in range(data_start_row, sheet1.max_row + 2):
            if r_idx > sheet1.max_row:
                data_end_row = sheet1.max_row
                break
            
            cell_value = sheet1.cell(row=r_idx, column=1).value
            if cell_value is None or (isinstance(cell_value, str) and footer_pattern.match(cell_value)):
                data_end_row = r_idx - 1
                break
            
            if r_idx == sheet1.max_row:
                data_end_row = r_idx
                break
                
        if not data_end_row or data_end_row < data_start_row:
            raise ValueError("Could not determine the data range after finding the header.")

        # Populate lookup dictionary from the source data
        data1_lookup = {}
        for row_num in range(data_start_row, data_end_row + 1):
            raw_name = sheet1.cell(row=row_num, column=1).value
            value_str = sheet1.cell(row=row_num, column=2).value
            
            if raw_name:
                normalized_key = normalize_d1_name(str(raw_name))
                numeric_value = None
                if isinstance(value_str, (int, float)):
                    numeric_value = value_str
                elif isinstance(value_str, str):
                    try:
                        numeric_value = float(value_str.replace(',', ''))
                    except (ValueError, TypeError):
                        pass

                # *** CORRECTED LOGIC HERE ***
                if normalized_key in SUMMABLE_NORMALIZED_KEYS:
                    if numeric_value is not None:
                        current_total = data1_lookup.get(normalized_key) or 0
                        data1_lookup[normalized_key] = current_total + numeric_value
                else: # For non-summable keys, only add if it's not already there
                    if normalized_key not in data1_lookup:
                        data1_lookup[normalized_key] = numeric_value

        # Prepare the new dataset based on the mapping
        dataset2_raw_output = []
        for item_map in MAP_D2_TO_D1:
            d2_name = item_map["d2_display_name"]
            raw_value = None
            
            if item_map["type"] == "direct":
                lookup_key = item_map["d1_lookup_names"][0] if item_map["d1_lookup_names"] else None
                if lookup_key:
                    raw_value = data1_lookup.get(lookup_key)
            elif item_map["type"] == "calculation" and item_map["op"] == "sum":
                current_sum = 0
                has_value = False
                for lookup_key in item_map["d1_lookup_names"]:
                    val = data1_lookup.get(lookup_key)
                    if isinstance(val, (int, float)):
                        current_sum += val
                        has_value = True
                raw_value = current_sum if has_value else None
            
            dataset2_raw_output.append([d2_name, raw_value])
        
        # --- Write "Dataset 2 Output" sheet ---
        number_format = '#,##0;"- "#,##0;0'
        output_sheet_name = "Dataset 2 Output"
        if output_sheet_name in workbook.sheetnames:
            del workbook[output_sheet_name]
        sheet2 = workbook.create_sheet(output_sheet_name)
        
        for r_idx, (name, val) in enumerate(dataset2_raw_output, 1):
            sheet2.cell(row=r_idx, column=1, value=name)
            cell_b = sheet2.cell(row=r_idx, column=2)
            if isinstance(val, (int, float)):
                cell_b.value = val
                cell_b.number_format = number_format
        
        sheet2.column_dimensions['A'].width = 40
        sheet2.column_dimensions['B'].width = 20

        # --- Write "Filtered Output" sheet ---
        filtered_sheet_name = "Filtered Output"
        if filtered_sheet_name in workbook.sheetnames:
            del workbook[filtered_sheet_name]
        sheet3 = workbook.create_sheet(filtered_sheet_name)
        
        filtered_row = 1
        for name, val in dataset2_raw_output:
            if isinstance(val, (int, float)):
                sheet3.cell(row=filtered_row, column=1, value=name)
                cell_b_filtered = sheet3.cell(row=filtered_row, column=2)
                cell_b_filtered.value = val
                cell_b_filtered.number_format = number_format
                filtered_row += 1
                
        sheet3.column_dimensions['A'].width = 40
        sheet3.column_dimensions['B'].width = 20

        # Save the modified workbook to an in-memory buffer
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    except Exception as e:
        # Raise a more informative exception
        raise Exception(f"An error occurred in pl_converter: {e}")

