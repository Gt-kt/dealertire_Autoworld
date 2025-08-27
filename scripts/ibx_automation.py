# /scripts/ibx_automation.py
import pandas as pd
import openpyxl
import numpy as np
import io

# --- All Configuration Constants (Copied from original script) ---
# B2B Config
CONFIRMED_STATUS_B2B = ['확정', '준비', '완료', '배송', '입금']
TIRE_BRANDS_B2B = ['피렐리', '금호', '한국', '넥센', '라우펜', '콘티넨탈', '브리지스톤', '미쉐린', '굿이어', '요코하마', '던롭', '프레데터', '쿠퍼']
BATTERY_BRANDS_B2B = ['아트라스 BX', '로케트배터리', '델코배터리', '바르타배터리', '한국배터리']
ENGINE_OIL_BRANDS_B2B = ['Kixx', 'ROWE', '캐스트롤']
VALVE_BRANDS_B2B = ['밸브']
BALANCE_WEIGHT_BRANDS_B2B = ['밸런스납']
WASTE_TIRE_BRANDS_B2B = ['폐타이어 수거 이용권']
B2B_INPUT_REQ_COLS = ['타이어가격', '수량', '상태', 'Brand']
VALUE_COLS_TO_CHECK_AND_AGGREGATE_B2B = ['배송비', '수량', '상품쿠폰', '배송비쿠폰', '포인트', '상품별 영업할인', '직원할인', '정산금액']
TIRE_OUTPUT_COLUMN_MAPPING_B2B = {
    '수량': 'E',
    '상품가': 'F',
    '배송비': 'G',
    '쿠폰': 'H',
    '포인트': 'I',
    '상품별 영업할인': 'J',
    '직원할인': 'L',
    '정산금액': 'M',
}
TIRE_OUTPUT_BRAND_COLUMN_B2B = 'D'
TIRE_OUTPUT_START_ROW_B2B = 4
TIRE_OUTPUT_END_ROW_B2B = 16
OTHER_CATEGORY_ROW_MAPPING_B2B = {'배터리': 19, '엔진오일': 20, '밸런스납': 21, '밸브': 22, '폐타이어': 23}
OTHER_CATEGORY_NAME_COLUMN_B2B = 'D'

# B2C Config
CONFIRMED_STATUS_B2C = ['확정', '배송', '완료', '입금', '준비']
TIRE_BRANDS_B2C = TIRE_BRANDS_B2B
BATTERY_BRANDS_B2C = BATTERY_BRANDS_B2B
B2C_INPUT_REQ_COLS = ['상태', 'Brand', 'Part No', '수량']
VALUE_COLS_TO_CHECK_AND_AGGREGATE_B2C = ['배송비', '수량', '상품쿠폰', '배송비쿠폰', '포인트', '정산금액']
TIRE_OUTPUT_COLUMN_MAPPING_B2C = {'수량': 'E', '상품가': 'F', '배송비': 'G', '쿠폰': 'H', '포인트': 'I', '정산금액': 'L'}
TIRE_OUTPUT_BRAND_COLUMN_B2C = 'D'
TIRE_OUTPUT_START_ROW_B2C = 4
TIRE_OUTPUT_END_ROW_B2C = 14
OTHER_CATEGORY_ROW_MAPPING_B2C = {'배터리': 17, '기타상품': 18, '용역': 21}
OTHER_CATEGORY_NAME_COLUMN_B2C = 'D'


def load_and_prepare_first_file(file_stream, data_type):
    df = pd.read_excel(file_stream, engine='openpyxl')
    
    if data_type == 'b2b':
        required_initial_cols = B2B_INPUT_REQ_COLS
        value_cols_to_check = VALUE_COLS_TO_CHECK_AND_AGGREGATE_B2B
    else: # b2c
        required_initial_cols = B2C_INPUT_REQ_COLS
        if '타이어가격' not in df.columns and '상품가' not in df.columns:
            raise ValueError("B2C 데이터에 '상품가'를 계산할 '타이어가격' 또는 '상품가' 컬럼이 없습니다.")
        if '타이어가격' in df.columns and '타이어가격' not in required_initial_cols:
            required_initial_cols.append('타이어가격')
        value_cols_to_check = VALUE_COLS_TO_CHECK_AND_AGGREGATE_B2C

    missing_cols = [col for col in required_initial_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"입력 파일에 필수 컬럼이 없습니다: {', '.join(missing_cols)}")

    if '상품가' not in df.columns:
        if '타이어가격' not in df.columns:
            raise KeyError("B2C '상품가' 계산에 필요한 '타이어가격' 컬럼이 없습니다.")
        df['타이어가격'] = pd.to_numeric(df['타이어가격'], errors='coerce').fillna(0)
        df['수량'] = pd.to_numeric(df['수량'], errors='coerce').fillna(0)
        df['상품가'] = df['타이어가격'] * df['수량']
    else:
        df['상품가'] = pd.to_numeric(df['상품가'], errors='coerce').fillna(0)
        if '수량' in df.columns:
            df['수량'] = pd.to_numeric(df['수량'], errors='coerce').fillna(0)

    confirmed_status_list = CONFIRMED_STATUS_B2B if data_type == 'b2b' else CONFIRMED_STATUS_B2C
    df['상태'] = df['상태'].astype(str).str.strip()
    df_filtered = df[df['상태'].notna() & df['상태'].isin(confirmed_status_list)].copy()

    if df_filtered.empty:
        return None

    all_value_cols = sorted(list(set(['상품가'] + value_cols_to_check)))
    for col in all_value_cols:
        if col not in df_filtered.columns:
            df_filtered[col] = 0
        else:
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce').fillna(0)
            
    return df_filtered


def aggregate_data(df_filtered, data_type):
    if df_filtered is None or df_filtered.empty:
        return None, None

    if data_type == 'b2b':
        # B2B Aggregation Logic
        cols_to_agg = ['상품가'] + [col for col in VALUE_COLS_TO_CHECK_AND_AGGREGATE_B2B if col in df_filtered.columns]
        df_filtered['Brand'] = df_filtered['Brand'].fillna('알수없음').astype(str).str.strip()
        aggregated_by_brand = df_filtered.groupby('Brand', dropna=False)[cols_to_agg].sum().reset_index()
        aggregated_by_brand['쿠폰'] = aggregated_by_brand.get('상품쿠폰', 0) + aggregated_by_brand.get('배송비쿠폰', 0)
        
        def categorize_brand_b2b(brand):
            brand_str = str(brand).strip()
            if brand_str in TIRE_BRANDS_B2B: return 'Tire'
            if brand_str in BATTERY_BRANDS_B2B: return '배터리'
            if brand_str in ENGINE_OIL_BRANDS_B2B: return '엔진오일'
            if brand_str in VALVE_BRANDS_B2B: return '밸브'
            if brand_str in BALANCE_WEIGHT_BRANDS_B2B: return '밸런스납'
            if brand_str in WASTE_TIRE_BRANDS_B2B: return '폐타이어'
            return '기타상품'
            
        aggregated_by_brand['Category'] = aggregated_by_brand['Brand'].apply(categorize_brand_b2b)
        tire_data_agg = aggregated_by_brand[aggregated_by_brand['Category'] == 'Tire'].copy()
        other_data_agg = aggregated_by_brand[aggregated_by_brand['Category'] != 'Tire'].copy()
        
        other_category_summary = pd.DataFrame()
        if not other_data_agg.empty:
            cols_for_category_agg = [col for col in aggregated_by_brand.columns if col not in ['Brand', 'Category', '상품쿠폰', '배송비쿠폰']]
            other_category_summary = other_data_agg.groupby('Category')[cols_for_category_agg].sum().reset_index()
    
    else: # b2c
        # B2C Aggregation Logic
        cols_to_agg = sorted(list(set(['상품가'] + [col for col in VALUE_COLS_TO_CHECK_AND_AGGREGATE_B2C if col in df_filtered.columns])))
        if 'Part No' not in df_filtered.columns or 'Brand' not in df_filtered.columns:
            raise ValueError("B2C 처리에 'Part No' 또는 'Brand' 컬럼이 필요합니다.")
        
        def categorize_b2c_item(row):
            if str(row.get('Part No', '')).strip().upper().startswith('B'): return '용역'
            if str(row.get('Brand', '')).strip() in BATTERY_BRANDS_B2C: return '배터리'
            if str(row.get('Brand', '')).strip() in TIRE_BRANDS_B2C: return 'Tire'
            return '기타상품'
            
        df_filtered['Category'] = df_filtered.apply(categorize_b2c_item, axis=1)
        
        tire_data_raw = df_filtered[df_filtered['Category'] == 'Tire']
        tire_data_agg = tire_data_raw.groupby('Brand')[cols_to_agg].sum().reset_index() if not tire_data_raw.empty else pd.DataFrame()

        other_categories_raw = df_filtered[df_filtered['Category'] != 'Tire']
        other_category_summary = other_categories_raw.groupby('Category')[cols_to_agg].sum().reset_index() if not other_categories_raw.empty else pd.DataFrame()

        for df in [tire_data_agg, other_category_summary]:
            if not df.empty:
                df['쿠폰'] = df.get('상품쿠폰', 0) + df.get('배송비쿠폰', 0)
    
    # Final processing for both B2B and B2C
    cols_to_divide = ['상품가', '배송비', '쿠폰', '포인트', '상품별 영업할인', '직원할인', '정산금액']
    for df in [tire_data_agg, other_category_summary]:
        if not df.empty:
            for col in cols_to_divide:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0) / 1.1
    
    output_col_mapping = TIRE_OUTPUT_COLUMN_MAPPING_B2B if data_type == 'b2b' else TIRE_OUTPUT_COLUMN_MAPPING_B2C
    
    tire_data_final = pd.DataFrame()
    if not tire_data_agg.empty:
        tire_data_final = tire_data_agg[['Brand'] + list(output_col_mapping.keys())]

    other_category_final = pd.DataFrame()
    if not other_category_summary.empty:
        other_category_final = other_category_summary[['Category'] + list(output_col_mapping.keys())]
        
    return tire_data_final, other_category_final


def update_template_file(template_stream, tire_data, other_data, data_type, sheet_name):
    # This function replaces the xlwings logic with openpyxl
    wb = openpyxl.load_workbook(template_stream)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}'를 템플릿 파일에서 찾을 수 없습니다. 사용 가능한 시트: {', '.join(wb.sheetnames)}")
    sheet = wb[sheet_name]
    
    # Get Type-Specific Configurations
    if data_type == 'b2b':
        tire_col, tire_start, tire_end, tire_map = TIRE_OUTPUT_BRAND_COLUMN_B2B, TIRE_OUTPUT_START_ROW_B2B, TIRE_OUTPUT_END_ROW_B2B, TIRE_OUTPUT_COLUMN_MAPPING_B2B
        other_col, other_map, other_row_map = OTHER_CATEGORY_NAME_COLUMN_B2B, TIRE_OUTPUT_COLUMN_MAPPING_B2B, OTHER_CATEGORY_ROW_MAPPING_B2B
    else: # b2c
        tire_col, tire_start, tire_end, tire_map = TIRE_OUTPUT_BRAND_COLUMN_B2C, TIRE_OUTPUT_START_ROW_B2C, TIRE_OUTPUT_END_ROW_B2C, TIRE_OUTPUT_COLUMN_MAPPING_B2C
        other_col, other_map, other_row_map = OTHER_CATEGORY_NAME_COLUMN_B2C, TIRE_OUTPUT_COLUMN_MAPPING_B2C, OTHER_CATEGORY_ROW_MAPPING_B2C

    # Update Tire Data
    if tire_data is not None and not tire_data.empty:
        tire_data_dict = {row['Brand']: row for _, row in tire_data.iterrows()}
        for row_num in range(tire_start, tire_end + 1):
            brand_in_sheet = sheet[f"{tire_col}{row_num}"].value
            if brand_in_sheet and brand_in_sheet.strip() in tire_data_dict:
                data_row = tire_data_dict[brand_in_sheet.strip()]
                for data_col_name, sheet_col_letter in tire_map.items():
                    value = data_row.get(data_col_name, 0)
                    sheet[f"{sheet_col_letter}{row_num}"].value = float(value) if pd.notna(value) else 0

    # Update Other Category Data
    if other_data is not None and not other_data.empty:
        other_data_dict = {row['Category']: row for _, row in other_data.iterrows()}
        for category_name, row_num in other_row_map.items():
            if category_name in other_data_dict:
                data_row = other_data_dict[category_name]
                for data_col_name, sheet_col_letter in other_map.items():
                     value = data_row.get(data_col_name, 0)
                     sheet[f"{sheet_col_letter}{row_num}"].value = float(value) if pd.notna(value) else 0

    # Save to memory buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


def process_files(data_type, sheet_name, input_file, template_file):
    """Main function to orchestrate the processing."""
    df_prepared = load_and_prepare_first_file(input_file, data_type)
    if df_prepared is None:
        raise ValueError(f"{data_type.upper()} 유형의 처리할 데이터가 없습니다.")

    processed_tire_data, processed_other_data = aggregate_data(df_prepared, data_type)
    if (processed_tire_data is None or processed_tire_data.empty) and \
       (processed_other_data is None or processed_other_data.empty):
        raise ValueError("집계 후 업데이트할 데이터가 없습니다.")
        
    return update_template_file(template_file, processed_tire_data, processed_other_data, data_type, sheet_name)
