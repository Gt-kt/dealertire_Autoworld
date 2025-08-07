# /scripts/margin_by_tire.py
import pandas as pd
import numpy as np
import io
from openpyxl.styles import Font

# -- Data Mappings --
BLACK_CIRCLE_MAP = {
    '1138168227': 'AJ제휴', 'cardoc': '카닥', '2208843430': '한국타이어 제휴',
    'halla': 'HL-유통', '1258130627': 'HL-퀀텀', '7538102566': 'HL-퀀텀',
    '1168136248': '현대캐피탈', 'kbcar': 'KB차차차', 'master': '마스터',
    'tscom': '티스테이션닷컴', 'HLFLEETON': '플릿온', 'cardocmall': '카닥몰',
    '5248800237': '테슬라(Tesla)', 'coupang2': '쿠팡', 'coupang': '쿠팡', 'coupang1': '쿠팡'
}

TIREPICK_MAP = {
    '7988101842': '타이어픽', 'TIREPICK': '타이어픽'
}

# --- Helper Functions for Data Processing ---

def create_new_columns(df):
    """Adds new columns and adjusts for VAT."""
    for col in ['수량', '타이어가격', '정산금액', '판매금액']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    df['정산금액'] = df['정산금액'] / 1.1
    df['판매금액'] = df['판매금액'] / 1.1
    df['상품가'] = (df['수량'] * df['타이어가격']) / 1.1

    df['블랙서클'] = df['주문ID'].astype(str).map(BLACK_CIRCLE_MAP)
    df['타이어픽'] = df['주문ID'].astype(str).map(TIREPICK_MAP)
    return df

def add_calculations_and_sort(pivot_table):
    """Adds calculated columns, rounds values, sorts, and reorders the pivot table."""
    if pivot_table.empty:
        return pivot_table

    total_row = pivot_table.loc[['총합계']].copy() if '총합계' in pivot_table.index else None
    pivot_table = pivot_table.drop('총합계', errors='ignore')

    with np.errstate(divide='ignore', invalid='ignore'):
        pivot_table['개당매입가'] = (pivot_table['정산금액'] / pivot_table['수량']).replace([np.inf, -np.inf], 0).fillna(0)
        pivot_table['개당판매가'] = (pivot_table['판매금액'] / pivot_table['수량']).replace([np.inf, -np.inf], 0).fillna(0)
        pivot_table['개당마진'] = pivot_table['개당판매가'] - pivot_table['개당매입가']
        pivot_table['총마진'] = pivot_table['개당마진'] * pivot_table['수량']
        pivot_table['마진율'] = (pivot_table['총마진'] / pivot_table['판매금액']).replace([np.inf, -np.inf], 0).fillna(0)
    
    pivot_table = pivot_table.sort_values(by='총마진', ascending=False)
    
    if total_row is not None:
        with np.errstate(divide='ignore', invalid='ignore'):
            total_row['개당매입가'] = (total_row['정산금액'] / total_row['수량']).replace([np.inf, -np.inf], 0).fillna(0)
            total_row['개당판매가'] = (total_row['판매금액'] / total_row['수량']).replace([np.inf, -np.inf], 0).fillna(0)
            total_row['개당마진'] = total_row['개당판매가'] - total_row['개당매입가']
            total_row['총마진'] = total_row['개당마진'] * total_row['수량']
            total_row['마진율'] = (total_row['총마진'] / total_row['판매금액']).replace([np.inf, -np.inf], 0).fillna(0)

    for col in ['정산금액', '판매금액', '총마진']:
        if col in pivot_table.columns: pivot_table[col] = pivot_table[col].round(-3).astype(np.int64)
        if total_row is not None and col in total_row.columns: total_row[col] = total_row[col].round(-3).astype(np.int64)

    for col in ['상품가', '개당매입가', '개당판매가', '개당마진']:
        if col in pivot_table.columns: pivot_table[col] = pivot_table[col].round().astype(np.int64)
        if total_row is not None and col in total_row.columns: total_row[col] = total_row[col].round().astype(np.int64)
            
    if total_row is not None:
        pivot_table = pd.concat([pivot_table, total_row])

    final_columns = ['수량', '정산금액', '개당매입가', '판매금액', '개당판매가', '개당마진', '총마진', '마진율']
    for col in final_columns:
        if col not in pivot_table.columns: pivot_table[col] = 0
    return pivot_table[final_columns]

def create_pivot_tables(df):
    """Creates the four required pivot tables."""
    status_filter = ['배송', '완료', '입금', '확정', '준비']
    brand_filter = [
        '피렐리', '금호', '한국', '넥센', '라우펜', '콘티넨탈',
        '브리지스톤', '미쉐린', '굿이어', '요코하마', '던롭', '프레데터', '쿠퍼'
    ]
    values_to_agg = ['수량', '정산금액', '상품가', '판매금액']
    
    filtered_df = df[df['상태'].isin(status_filter) & df['Brand'].isin(brand_filter)]

    pivot1 = pd.pivot_table(filtered_df, values=values_to_agg, index='Brand', aggfunc='sum')
    if not pivot1.empty: pivot1.loc['총합계'] = pivot1.sum()
    pivot1 = add_calculations_and_sort(pivot1)

    df_b2b = filtered_df[filtered_df['타이어픽'].isna()]
    pivot2 = pd.pivot_table(df_b2b, values=values_to_agg, index='Brand', aggfunc='sum')
    if not pivot2.empty: pivot2.loc['총합계'] = pivot2.sum()
    pivot2 = add_calculations_and_sort(pivot2)

    df_tirepick = filtered_df[filtered_df['타이어픽'] == '타이어픽']
    pivot3 = pd.pivot_table(df_tirepick, values=values_to_agg, index='Brand', aggfunc='sum')
    if not pivot3.empty: pivot3.loc['총합계'] = pivot3.sum()
    pivot3 = add_calculations_and_sort(pivot3)

    df_b2b_channels = df[df['블랙서클'].notna()]
    pivot4 = pd.pivot_table(df_b2b_channels, values=values_to_agg, index='블랙서클', aggfunc='sum')
    if not pivot4.empty: pivot4.loc['총합계'] = pivot4.sum()
    pivot4 = add_calculations_and_sort(pivot4)

    return pivot1, pivot2, pivot3, pivot4

def apply_number_formats(sheet, pivot_table, start_row):
    """Applies number formatting to the Excel sheet."""
    if pivot_table.empty: return
    
    int_format, percent_format = '#,##0', '0.0%'
    try:
        margin_rate_col_idx = list(pivot_table.columns).index('마진율') + 2
    except ValueError:
        margin_rate_col_idx = -1

    for r_idx in range(start_row + 2, start_row + 2 + len(pivot_table)):
        for c_idx in range(2, len(pivot_table.columns) + 2):
            cell = sheet.cell(row=r_idx, column=c_idx)
            if isinstance(cell.value, (int, float, np.number)):
                cell.number_format = percent_format if cell.column == margin_rate_col_idx else int_format

def save_to_excel(pivot1, pivot2, pivot3, pivot4, date_range):
    """Saves the pivot tables to a new Excel file in memory."""
    output_buffer = io.BytesIO()
    bold_font = Font(name='Calibri', size=12, bold=True)

    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        sheet_name_margin, sheet_name_b2b = 'Item별 마진', 'B2B 채널별'
        
        start_row_p1 = 2
        start_row_p2 = start_row_p1 + len(pivot1) + 4 if not pivot1.empty else start_row_p1
        start_row_p3 = start_row_p2 + len(pivot2) + 4 if not pivot2.empty else start_row_p2
        start_row_p4 = 2

        if not pivot1.empty: pivot1.to_excel(writer, sheet_name=sheet_name_margin, startrow=start_row_p1)
        if not pivot2.empty: pivot2.to_excel(writer, sheet_name=sheet_name_margin, startrow=start_row_p2)
        if not pivot3.empty: pivot3.to_excel(writer, sheet_name=sheet_name_margin, startrow=start_row_p3)
        if not pivot4.empty: pivot4.to_excel(writer, sheet_name=sheet_name_b2b, startrow=start_row_p4)

        sheet1 = writer.sheets[sheet_name_margin]
        sheet2 = writer.sheets[sheet_name_b2b]
        
        sheet1.cell(row=1, column=1, value=date_range).font = bold_font
        sheet2.cell(row=1, column=1, value=date_range).font = bold_font

        sheet1.cell(row=start_row_p1 + 1, column=1, value="1. 전체 (Total)").font = bold_font
        apply_number_formats(sheet1, pivot1, start_row_p1)
        sheet1.cell(row=start_row_p2 + 1, column=1, value="2. 블랙서클 (Blackcircle)").font = bold_font
        apply_number_formats(sheet1, pivot2, start_row_p2)
        sheet1.cell(row=start_row_p3 + 1, column=1, value="3. 타이어픽 (Tire-pick)").font = bold_font
        apply_number_formats(sheet1, pivot3, start_row_p3)
        sheet2.cell(row=start_row_p4 + 1, column=1, value="4. B2B 채널별 (B2B by Channel)").font = bold_font
        apply_number_formats(sheet2, pivot4, start_row_p4)

    output_buffer.seek(0)
    return output_buffer

# --- Main Function to be Called by the Web App ---
def process_file(file_stream):
    """Handles the entire process for a single uploaded file."""
    try:
        df = pd.read_excel(file_stream)
        
        date_range_str = "기간 정보를 가져올 수 없습니다."
        if '주문일자' in df.columns:
            try:
                df['주문일자'] = pd.to_datetime(df['주문일자'], errors='coerce')
                df.dropna(subset=['주문일자'], inplace=True)
                min_date = df['주문일자'].min().strftime('%Y-%m-%d')
                max_date = df['주문일자'].max().strftime('%Y-%m-%d')
                date_range_str = f"기간 (Period): {min_date} ~ {max_date}"
            except Exception:
                pass # Ignore errors in date parsing for now

        df = create_new_columns(df)
        pivot1, pivot2, pivot3, pivot4 = create_pivot_tables(df)
        return save_to_excel(pivot1, pivot2, pivot3, pivot4, date_range_str)

    except Exception as e:
        raise ValueError(f"An error occurred during processing: {e}")
